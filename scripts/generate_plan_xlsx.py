#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
考研规划报告 Excel 生成器
==========================
读取用户数据 JSON，生成多 sheet 结构化 Excel 计划表。

用法:
    python generate_plan_xlsx.py --data user_data.json --output 计划表.xlsx
    python generate_plan_xlsx.py --data user_data.json                  # 默认输出

依赖: openpyxl (pip install openpyxl)
"""

import json
import os
import sys
import argparse
from datetime import datetime, timedelta
from pathlib import Path

# 修复 Windows GBK 控制台编码问题
if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
        numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table
except ImportError:
    print("❌ 缺少 openpyxl 库。请运行: pip install openpyxl")
    sys.exit(1)


# ============================================================
# 样式常量
# ============================================================

# 颜色
PRIMARY = "4f46e5"
PRIMARY_LIGHT = "e0e7ff"
DANGER = "ef4444"
DANGER_LIGHT = "fef2f2"
WARNING = "f59e0b"
WARNING_LIGHT = "fffbeb"
SUCCESS = "10b981"
SUCCESS_LIGHT = "ecfdf5"
DARK = "1e293b"
GRAY = "64748b"
WHITE = "ffffff"
LIGHT_GRAY = "f8fafc"
BORDER_COLOR = "e2e8f0"

# 字体
FONT_TITLE = Font(name="微软雅黑", size=18, bold=True, color=DARK)
FONT_SECTION = Font(name="微软雅黑", size=14, bold=True, color=PRIMARY)
FONT_HEADER = Font(name="微软雅黑", size=11, bold=True, color=WHITE)
FONT_NORMAL = Font(name="微软雅黑", size=11, color=DARK)
FONT_BOLD = Font(name="微软雅黑", size=11, bold=True, color=DARK)
FONT_SMALL = Font(name="微软雅黑", size=10, color=GRAY)
FONT_DANGER = Font(name="微软雅黑", size=11, bold=True, color=DANGER)
FONT_SUCCESS = Font(name="微软雅黑", size=11, bold=True, color=SUCCESS)
FONT_WARNING = Font(name="微软雅黑", size=11, bold=True, color=WARNING)

# 填充
FILL_HEADER = PatternFill(start_color=PRIMARY, end_color=PRIMARY, fill_type="solid")
FILL_HEADER_DANGER = PatternFill(start_color=DANGER, end_color=DANGER, fill_type="solid")
FILL_LIGHT = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
FILL_PRIMARY_LIGHT = PatternFill(start_color=PRIMARY_LIGHT, end_color=PRIMARY_LIGHT, fill_type="solid")
FILL_DANGER_LIGHT = PatternFill(start_color=DANGER_LIGHT, end_color=DANGER_LIGHT, fill_type="solid")
FILL_WARNING_LIGHT = PatternFill(start_color=WARNING_LIGHT, end_color=WARNING_LIGHT, fill_type="solid")
FILL_SUCCESS_LIGHT = PatternFill(start_color=SUCCESS_LIGHT, end_color=SUCCESS_LIGHT, fill_type="solid")
FILL_WHITE = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

# 对齐
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

# 边框
THIN_BORDER = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)


# ============================================================
# 工具函数
# ============================================================

def apply_header_style(ws, row, col_start, col_end, fill=FILL_HEADER, font=FONT_HEADER):
    """为表头行应用样式。"""
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.fill = fill
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER


def apply_row_style(ws, row, col_start, col_end, font=FONT_NORMAL, fill=None, alignment=ALIGN_LEFT):
    """为数据行应用样式。"""
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.alignment = alignment
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill


def set_col_widths(ws, widths: list[tuple[int, int]]):
    """设置列宽。widths: [(col_index, width), ...]"""
    for col, width in widths:
        ws.column_dimensions[get_column_letter(col)].width = width


def write_title(ws, row, col, text, merge_end_col=None):
    """写入标题。"""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = FONT_SECTION
    cell.alignment = ALIGN_LEFT
    if merge_end_col:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end_col)


def write_section_title(ws, row, col, text, merge_end_col=None):
    """写入小节标题。"""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(name="微软雅黑", size=12, bold=True, color=DARK)
    cell.alignment = ALIGN_LEFT
    if merge_end_col:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_end_col)


# ============================================================
# Sheet 生成函数
# ============================================================

def create_overview_sheet(wb: Workbook, data: dict):
    """Sheet 1: 综合概览"""
    ws = wb.active
    ws.title = "📋 综合概览"

    student = data.get("student", {})
    assessment = data.get("assessment", {})
    subjects = data.get("subjects", {})

    # 列宽
    set_col_widths(ws, [(1, 22), (2, 35), (3, 22), (4, 35), (5, 22), (6, 35)])

    # 大标题
    ws.merge_cells("A1:F1")
    title_cell = ws.cell(row=1, column=1, value=f"🎓 考研全程规划报告 — {student.get('name', '同学')}")
    title_cell.font = FONT_TITLE
    title_cell.alignment = ALIGN_CENTER
    ws.row_dimensions[1].height = 40

    # 基本信息
    row = 3
    write_title(ws, row, 1, "👤 基本信息")
    row += 1
    info_fields = [
        ("姓名", "name"), ("本科院校", "undergrad_school"), ("本科专业", "undergrad_major"),
        ("目标院校", "target_school"), ("目标专业", "target_major"),
        ("目标总分", "target_total_score"), ("英语水平", "english_level"),
        ("数学基础", "math_level"), ("每日可用时间", "daily_hours"),
        ("是否跨专业", "is_cross_major"), ("是否在职", "is_working"),
        ("考试日期", "exam_date"),
    ]
    headers = ["项目", "内容", "项目", "内容"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    apply_header_style(ws, row, 1, 4)

    row += 1
    for i, (label, key) in enumerate(info_fields):
        col_offset = (i % 2) * 3
        if i % 2 == 0 and i > 0:
            row += 1
        val = student.get(key, "—")
        if isinstance(val, bool):
            val = "是" if val else "否"
        ws.cell(row=row, column=1 + col_offset, value=label).font = FONT_BOLD
        ws.cell(row=row, column=1 + col_offset).alignment = ALIGN_LEFT
        ws.cell(row=row, column=1 + col_offset).border = THIN_BORDER
        ws.cell(row=row, column=1 + col_offset).fill = FILL_LIGHT
        ws.cell(row=row, column=2 + col_offset, value=str(val)).alignment = ALIGN_LEFT
        ws.cell(row=row, column=2 + col_offset).border = THIN_BORDER

    # 评估结论
    row += 2
    write_title(ws, row, 1, "📊 评估结论")
    row += 1
    eval_data = [
        ("加权总分", f"{assessment.get('weighted_score', '—')} / 5.0"),
        ("评估等级", assessment.get("level", "—")),
        ("自评上岸率", f"{assessment.get('probability', '—')}%"),
    ]
    for i, (label, value) in enumerate(eval_data):
        ws.cell(row=row + i, column=1, value=label).font = FONT_BOLD
        ws.cell(row=row + i, column=1).border = THIN_BORDER
        ws.cell(row=row + i, column=1).fill = FILL_LIGHT
        ws.cell(row=row + i, column=2, value=value).border = THIN_BORDER

    # 各科进度
    row += 5
    write_title(ws, row, 1, "📈 各科进度与目标")
    row += 1
    subj_headers = ["科目", "当前进度", "目标分数", "状态"]
    for c, h in enumerate(subj_headers, 1):
        ws.cell(row=row, column=c, value=h)
    apply_header_style(ws, row, 1, 4)

    row += 1
    for name, info in subjects.items():
        pct = info.get("progress_pct", 0)
        target = info.get("target_score", "—")
        status = "🟢 正常" if pct >= 70 else ("🟡 跟进" if pct >= 40 else "🔴 优先")
        ws.cell(row=row, column=1, value=name).font = FONT_BOLD
        ws.cell(row=row, column=2, value=f"{pct}%")
        ws.cell(row=row, column=3, value=str(target))
        ws.cell(row=row, column=4, value=status)
        fill = FILL_SUCCESS_LIGHT if pct >= 70 else (FILL_WARNING_LIGHT if pct >= 40 else FILL_DANGER_LIGHT)
        apply_row_style(ws, row, 1, 4, fill=fill)
        row += 1

    # 生成信息
    row += 2
    ws.cell(row=row, column=1, value=f"报告生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M')}").font = FONT_SMALL


def create_school_sheet(wb: Workbook, data: dict):
    """Sheet 2: 择校分析"""
    ws = wb.create_sheet("🏫 择校分析")
    set_col_widths(ws, [(1, 20), (2, 12), (3, 18), (4, 16), (5, 14), (6, 14), (7, 35)])

    row = 1
    write_title(ws, row, 1, "🏫 三档院校推荐", 7)
    row += 1

    school_data = data.get("school_selection", {})
    headers = ["院校", "层次", "专业", "近年复试线", "报录比", "录取概率", "备注"]

    for tier_key, tier_label, emoji in [
        ("sprint", "冲刺档", "🚀"), ("stable", "稳妥档", "🎯"), ("safety", "保底档", "🛡️")
    ]:
        schools = school_data.get(tier_key, [])
        if not schools:
            continue

        row += 1
        write_section_title(ws, row, 1, f"{emoji} {tier_label}", 7)
        row += 1

        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        fill = FILL_HEADER_DANGER if tier_key == "sprint" else (PatternFill(start_color=WARNING, end_color=WARNING, fill_type="solid") if tier_key == "stable" else FILL_HEADER)
        apply_header_style(ws, row, 1, 7, fill=fill)
        row += 1

        for s in schools:
            vals = [s.get("name", ""), s.get("level", ""), s.get("major", ""),
                    s.get("cutoff", ""), s.get("ratio", ""), s.get("probability", ""), s.get("note", "")]
            for c, v in enumerate(vals, 1):
                ws.cell(row=row, column=c, value=v)
            apply_row_style(ws, row, 1, 7)
            row += 1

    # 风险提示
    risk_notes = school_data.get("risk_notes", [])
    if risk_notes:
        row += 1
        write_section_title(ws, row, 1, "⚠️ 风险提示", 7)
        row += 1
        for note in risk_notes:
            ws.cell(row=row, column=1, value=note).font = FONT_DANGER
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            row += 1


def create_study_plan_sheet(wb: Workbook, data: dict):
    """Sheet 3: 学习计划"""
    ws = wb.create_sheet("📅 学习计划")
    set_col_widths(ws, [(1, 18), (2, 16), (3, 30), (4, 28), (5, 14), (6, 30)])

    row = 1
    write_title(ws, row, 1, "📅 三阶段学习计划", 6)

    study = data.get("study_plan", {})
    phases = study.get("phases", [])

    headers = ["科目", "阶段", "任务", "推荐资料", "每日用时", "完成标准"]

    for ph in phases:
        row += 2
        phase_name = ph.get("phase_name", "")
        period = ph.get("period", "")
        write_section_title(ws, row, 1, f"{phase_name}（{period}）", 6)

        row += 1
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        apply_header_style(ws, row, 1, 6)
        row += 1

        for subj in ph.get("subjects", []):
            vals = [
                subj.get("name", ""), phase_name, subj.get("task", ""),
                subj.get("material", ""), str(subj.get("daily_hours", "")),
                subj.get("completion_criteria", "")
            ]
            for c, v in enumerate(vals, 1):
                ws.cell(row=row, column=c, value=v)
            apply_row_style(ws, row, 1, 6)
            row += 1

    # 月度时间轴
    months = study.get("months", [])
    if months:
        row += 2
        write_section_title(ws, row, 1, "📆 月度规划时间轴", 6)
        row += 1
        m_headers = ["月份", "主题", "详细内容", "", "", ""]
        for c, h in enumerate(m_headers, 1):
            ws.cell(row=row, column=c, value=h)
        apply_header_style(ws, row, 1, 6)
        row += 1
        for m in months:
            ws.cell(row=row, column=1, value=m.get("month", "")).font = FONT_BOLD
            ws.cell(row=row, column=2, value=m.get("title", ""))
            ws.cell(row=row, column=3, value=m.get("detail", ""))
            ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
            apply_row_style(ws, row, 1, 6)
            row += 1


def create_daily_sheet(wb: Workbook, data: dict):
    """Sheet 4: 每日作息"""
    ws = wb.create_sheet("⏰ 每日作息")
    set_col_widths(ws, [(1, 16), (2, 40), (3, 16), (4, 12)])

    row = 1
    write_title(ws, row, 1, "⏰ 每日作息表", 4)

    daily = data.get("daily_schedule", {})
    slots = daily.get("slots", [])

    if slots:
        row += 2
        headers = ["时间", "任务", "类型", "强度"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        apply_header_style(ws, row, 1, 4)
        row += 1

        for s in slots:
            intensity = s.get("intensity", "中")
            vals = [s.get("time", ""), s.get("task", ""), s.get("type", ""), intensity]
            for c, v in enumerate(vals, 1):
                ws.cell(row=row, column=c, value=v)
            fill = FILL_DANGER_LIGHT if intensity == "高" else (FILL_WARNING_LIGHT if intensity == "中" else None)
            apply_row_style(ws, row, 1, 4, fill=fill)
            row += 1

    # 保底任务
    min_tasks = daily.get("minimum_tasks", [])
    if min_tasks:
        row += 2
        write_section_title(ws, row, 1, "🛡️ 每日保底任务", 4)
        row += 1
        t_headers = ["科目", "最低量", "用时", ""]
        for c, h in enumerate(t_headers, 1):
            ws.cell(row=row, column=c, value=h)
        apply_header_style(ws, row, 1, 4)
        row += 1
        for t in min_tasks:
            vals = [t.get("subject", ""), t.get("amount", ""), t.get("time", ""), ""]
            for c, v in enumerate(vals, 1):
                ws.cell(row=row, column=c, value=v)
            apply_row_style(ws, row, 1, 4)
            row += 1


def create_weekly_sheet(wb: Workbook, data: dict):
    """Sheet 5: 每周里程碑"""
    ws = wb.create_sheet("📈 每周里程碑")
    set_col_widths(ws, [(1, 22), (2, 38), (3, 38), (4, 38), (5, 38)])

    row = 1
    write_title(ws, row, 1, "📈 每周冲刺里程碑", 5)

    weeks = data.get("weekly_milestones", [])
    if weeks:
        row += 2
        headers = ["周次", "数学", "408专业课", "政治", "英语"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        apply_header_style(ws, row, 1, 5)
        row += 1

        for w in weeks:
            vals = [w.get("week", ""), w.get("math", ""), w.get("cs408", ""),
                    w.get("politics", ""), w.get("english", "")]
            for c, v in enumerate(vals, 1):
                ws.cell(row=row, column=c, value=v)
            apply_row_style(ws, row, 1, 5)
            row += 1


def create_mock_exam_sheet(wb: Workbook, data: dict):
    """Sheet 6: 模考评估"""
    ws = wb.create_sheet("🎯 模考评估")
    set_col_widths(ws, [(1, 18), (2, 16), (3, 16), (4, 16), (5, 30)])

    row = 1
    write_title(ws, row, 1, "🎯 模考成绩动态评估", 5)

    mock = data.get("mock_exam", {})
    scores = mock.get("scores", {})
    targets = mock.get("target_scores", {})

    if scores:
        row += 2
        write_section_title(ws, row, 1, "成绩对标", 5)
        row += 1
        headers = ["科目", "模考成绩", "目标分数", "差距", "判断"]
        for c, h in enumerate(headers, 1):
            ws.cell(row=row, column=c, value=h)
        apply_header_style(ws, row, 1, 5)
        row += 1

        for subj in ["数学", "408专业课", "政治", "英语", "总分"]:
            s = scores.get(subj, "—")
            t = targets.get(subj, "—")
            try:
                gap = f"{s - t:+d}" if isinstance(s, (int, float)) and isinstance(t, (int, float)) else "—"
            except TypeError:
                gap = "—"
            risk = "🔴 危险" if isinstance(s, (int, float)) and isinstance(t, (int, float)) and (t - s) >= 15 else \
                   ("🟡 警戒" if isinstance(s, (int, float)) and isinstance(t, (int, float)) and (t - s) >= 5 else "🟢 安全")
            vals = [subj, str(s), str(t), gap, risk]
            for c, v in enumerate(vals, 1):
                ws.cell(row=row, column=c, value=v)
            fill = FILL_DANGER_LIGHT if "危险" in risk else (FILL_WARNING_LIGHT if "警戒" in risk else FILL_SUCCESS_LIGHT)
            apply_row_style(ws, row, 1, 5, fill=fill, font=FONT_BOLD if subj == "总分" else FONT_NORMAL)
            row += 1

    # 可达性推演
    projection = mock.get("projection", [])
    if projection:
        row += 2
        write_section_title(ws, row, 1, "总分可达性推演", 5)
        row += 1
        p_headers = ["科目", "当前", "中性预期", "乐观上限", ""]
        for c, h in enumerate(p_headers, 1):
            ws.cell(row=row, column=c, value=h)
        apply_header_style(ws, row, 1, 5)
        row += 1
        for pr in projection:
            vals = [pr.get("subject", ""), str(pr.get("current", "")),
                    str(pr.get("expected", "")), str(pr.get("optimistic", "")), ""]
            for c, v in enumerate(vals, 1):
                ws.cell(row=row, column=c, value=v)
            apply_row_style(ws, row, 1, 5)
            row += 1

    # 方案对比
    plans = mock.get("plans", [])
    if plans:
        row += 2
        write_section_title(ws, row, 1, "调整方案对比", 5)
        row += 1
        plan_headers = ["方案", "目标院校", "层次", "差距", "成功率"]
        for c, h in enumerate(plan_headers, 1):
            ws.cell(row=row, column=c, value=h)
        apply_header_style(ws, row, 1, 5)
        row += 1
        for plan in plans:
            vals = [plan.get("label", ""), plan.get("school", ""), plan.get("level", ""),
                    plan.get("gap", ""), plan.get("success_rate", "")]
            for c, v in enumerate(vals, 1):
                ws.cell(row=row, column=c, value=v)
            apply_row_style(ws, row, 1, 5)
            row += 1


# ============================================================
# 主流程
# ============================================================

def generate_plan_xlsx(data: dict, output_path: str):
    """主入口：生成多 sheet Excel 计划表。"""
    wb = Workbook()

    # 创建各 sheet
    create_overview_sheet(wb, data)
    create_school_sheet(wb, data)
    create_study_plan_sheet(wb, data)
    create_daily_sheet(wb, data)
    create_weekly_sheet(wb, data)
    create_mock_exam_sheet(wb, data)

    # 冻结首行
    for ws in wb.worksheets:
        ws.freeze_panes = ws.cell(row=2, column=1)

    # 保存
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description="考研规划报告 Excel 生成器",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例:
  python generate_plan_xlsx.py --data user_data.json
  python generate_plan_xlsx.py --data user_data.json --output 我的计划表.xlsx
        """
    )
    parser.add_argument("--data", required=True, help="用户数据 JSON 文件路径")
    parser.add_argument("--output", "-o", default=None, help="输出 Excel 文件路径")
    args = parser.parse_args()

    # 读取数据
    try:
        with open(args.data, "r", encoding="utf-8") as f:
            data = json.load(f)
    except FileNotFoundError:
        print(f"❌ 数据文件不存在: {args.data}", file=sys.stderr)
        print(f"   请先运行 collect_user_info.py 采集数据。", file=sys.stderr)
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"❌ 读取用户数据失败：JSON 格式有误。", file=sys.stderr)
        print(f"   错误位置: 第 {e.lineno} 行，第 {e.colno} 列", file=sys.stderr)
        print(f"   请检查 JSON 语法或重新采集数据。", file=sys.stderr)
        sys.exit(1)
    except UnicodeDecodeError:
        print(f"❌ 文件编码错误，请确保使用 UTF-8 编码。", file=sys.stderr)
        sys.exit(1)

    # 确定输出路径
    if args.output:
        output_path = args.output
    else:
        student_name = data.get("student", {}).get("name", "同学")
        output_path = f"考研计划表_{student_name}.xlsx"

    output_path = os.path.abspath(output_path)

    # 生成
    try:
        generate_plan_xlsx(data, output_path)
    except Exception as e:
        print(f"❌ Excel 生成失败: {e}", file=sys.stderr)
        print(f"   请检查 openpyxl 是否已安装: pip install openpyxl", file=sys.stderr)
        print(f"   请检查 JSON 数据字段是否完整。", file=sys.stderr)
        sys.exit(1)

    print(f"✅ Excel 计划表已生成: {output_path}")
    print(f"📄 文件大小: {os.path.getsize(output_path) / 1024:.1f} KB")
    print(f"📊 包含 {len(data.get('study_plan', {}).get('phases', []))} 个阶段 + 择校/作息/里程碑/模考 共 6 个 Sheet")


if __name__ == "__main__":
    main()
